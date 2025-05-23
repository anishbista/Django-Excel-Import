import pandas as pd
from openpyxl import load_workbook
from django.db import transaction
from products.models import Product
from importer.models import ImportJob, ImportLog
from utils.validators import ProductValidator


class ExcelProductProcessor:
    def __init__(self, import_job_id):
        self.import_job = ImportJob.objects.get(id=import_job_id)
        self.chunk_size = 100

    def process(self):
        try:
            self.import_job.status = "processing"
            self.import_job.save()

            excel_file = self.import_job.file.path
            workbook = load_workbook(excel_file, read_only=True)
            sheet = workbook.active

            total_rows = 0
            success_count = 0
            warning_count = 0
            error_count = 0

            rows = sheet.iter_rows(min_row=2, values_only=True)
            chunk = []
            for row_num, row in enumerate(rows, start=2):
                chunk.append(row)
                if len(chunk) == self.chunk_size:
                    total_rows += len(chunk)
                    chunk_success, chunk_warning, chunk_error = self._process_chunk(
                        chunk, row_num
                    )
                    success_count += chunk_success
                    warning_count += chunk_warning
                    error_count += chunk_error
                    chunk = []

            if chunk:
                total_rows += len(chunk)
                chunk_success, chunk_warning, chunk_error = self._process_chunk(
                    chunk, row_num
                )
                success_count += chunk_success
                warning_count += chunk_warning
                error_count += chunk_error

            self.import_job.status = "completed"
            self.import_job.total_rows = total_rows
            self.import_job.success_count = success_count
            self.import_job.warning_count = warning_count
            self.import_job.error_count = error_count
            self.import_job.save()

        except Exception as e:
            self.import_job.status = "failed"
            self.import_job.save()
            raise

    def _process_chunk(self, chunk, last_row_num):
        success_count = 0
        warning_count = 0
        error_count = 0

        for index, row in enumerate(chunk):
            row_num = last_row_num - len(chunk) + index + 1
            row_data = dict(zip(self._get_headers(), row))

            errors = ProductValidator.validate_required_fields(row_data, row_num)
            warnings = ProductValidator.validate_warning_fields(row_data, row_num)
            price_warnings = ProductValidator.validate_price(row_data, row_num)
            warnings.extend(price_warnings)

            if errors:
                error_count += 1
                for error in errors:
                    self._log_error(row_num, error)
                continue

            try:
                with transaction.atomic():
                    product = self._create_product(row_data)
                    success_count += 1

                    for warning in warnings:
                        warning_count += 1
                        self._log_warning(row_num, warning)

            except Exception as e:
                error_count += 1
                self._log_error(row_num, str(e))

        return success_count, warning_count, error_count

    def _get_headers(self):
        return [
            "id",
            "title",
            "image_link",
            "description",
            "link",
            "price",
            "sale_price",
            "shipping",
            "item_group_id",
            "availability",
            "additional_image_link",
            "brand",
            "gtin",
            "gender",
            "google_product_category",
            "product_type",
            "material",
            "pattern",
            "color",
            "product_length",
            "product_width",
            "product_height",
            "product_weight",
            "size",
            "lifestyle_image_link",
            "max_handling_time",
            "is_bundle",
            "Model",
            "condition",
        ]

    def _create_product(self, row_data):
        def preprocess_price(price_str):

            if not price_str:
                return None
            try:
                return float(price_str.replace(",", "").split()[0])
            except (ValueError, AttributeError):
                return None

        product_data = {
            "sku": row_data["id"],
            "title": row_data["title"],
            "description": row_data["description"],
            "link": row_data["link"],
            "image_link": row_data["image_link"],
            "availability": row_data["availability"],
            "price": preprocess_price(row_data["price"]),
            "condition": row_data["condition"],
            "brand": row_data["brand"],
            "gtin": row_data["gtin"],
            "sale_price": preprocess_price(row_data.get("sale_price")),
            "item_group_id": row_data.get("item_group_id"),
            "google_product_category": row_data.get("google_product_category"),
            "product_type": row_data.get("product_type"),
            "size": row_data.get("size"),
            "color": row_data.get("color"),
            "material": row_data.get("material"),
            "pattern": row_data.get("pattern"),
            "gender": row_data.get("gender"),
            "model": row_data.get("Model"),
        }

        for key, value in product_data.items():
            if isinstance(value, str) and not value.strip():
                product_data[key] = None

        product, created = Product.objects.update_or_create(
            sku=product_data["sku"], defaults=product_data
        )
        return product

    def _log_error(self, row_num, message):
        ImportLog.objects.create(
            job=self.import_job, log_type="error", row_number=row_num, message=message
        )

    def _log_warning(self, row_num, message):
        ImportLog.objects.create(
            job=self.import_job, log_type="warning", row_number=row_num, message=message
        )
